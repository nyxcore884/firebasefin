import firebase_admin
from firebase_admin import credentials, storage
import openpyxl
import io
import sys

if not firebase_admin._apps:
    firebase_admin.initialize_app()

bucket = storage.bucket('studio-9381016045-4d625.firebasestorage.app')

# List files in ingestion folder
blobs = list(bucket.list_blobs(prefix='ingestion/'))
print("Files in storage:")
for blob in blobs:
    print(f"  - {blob.name}")

# Download the specific file
target_file = "ingestion/Book1-1769185683888.xlsx"
blob = bucket.blob(target_file)

if blob.exists():
    print(f"\nDownloading {target_file}...")
    content = blob.download_as_bytes()
    
    # Parse with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    
    print(f"\nSheet name: {ws.title}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    print("\n--- First 15 rows (raw) ---")
    for row_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
        # Clean up for display
        row_str = [str(c)[:20] if c else '' for c in row[:15]]
        print(f"Row {row_idx}: {row_str}")
    
    # Check for month headers
    print("\n--- Searching for month headers ---")
    month_terms = ['january', 'february', 'march', 'april', 'may', 'june', 
                   'july', 'august', 'september', 'october', 'november', 'december',
                   'jan', 'feb', 'mar', 'apr', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
    
    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
        for cell in row:
            if cell and str(cell).lower().strip() in month_terms:
                print(f"  Found '{cell}' at row {row_idx}")
                break
else:
    print(f"File {target_file} not found!")
